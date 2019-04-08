import openpyxl

# Creating object of internal Workbook
wb = openpyxl.Workbook()

# Gives active sheet name
print(wb.active.title)

# Creates object of active sheet
sh = wb.active

# Changed active sheet name
sh.title = "Edited sheet name"

# Prints sheets name
print(sh.title)

# Write to active sheet cell A4
sh["A4"].value = "Hello World"

# Create new sheet
wb.create_sheet("Testing")

#Create object of new sheet
sh2 = wb["Testing"]

# Write to new sheets A3 cell
sh2["A3"] = "34343334"

# Create third sheet
wb.create_sheet("This is to be removed")

sh3 = wb["This is to be removed"]

# Remove sheet
wb.remove(sh3)

# Save workbook to location
wb.save("D:\\Demo.xlsx")
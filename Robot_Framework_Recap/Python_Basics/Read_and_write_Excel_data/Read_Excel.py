import openpyxl

# Load workbook
wb = openpyxl.load_workbook("TestData.xlsx")

print(wb.sheetnames)
print("\nActive sheet is " + wb.active.title)

# Create object of any sheet on which you want to work
sh = wb["TestData"]
print("\nChanged active sheet to " + sh.title)

# Getting data from specific cell
print("\nIn cell A3 is data " + str(sh["A3"].value))

# Getting data from specific cell
# Syntax is sh.cell(row,column)
# c1Data = sh.cell(row=1,column=3)
c1Data = sh.cell(1,3)
print("\nIn cell C1 is data " + str(c1Data.value))

b2Data = sh.cell(2, 2)
print("\nSpecific location of data " + str(b2Data.row) + str(b2Data.column))

# Read all data in Excel
# Find out how many rows are having data
columns = sh.max_column
rows = sh.max_row

print("\nTotal size of Excel is columns " +str(columns) + " and rows " + str(rows) + "\n")

# Print all data in Excel
# rows+1 is needed because in range last value is excluded by default
# In range(1,10) return values from 1 to 9
for i in range(1, rows+1):
    for j in range(1, columns+1):
        allData = sh.cell(i,j)
        print(allData.value)
        
print("\n********************** Second approach to get data from Excel **********************************\n")

# Second approach to get all data without using range
for r in sh["A1":"C4"]:
    for c in r:
        print(c.value)